import openpyxl
import pyautogui
from time import sleep

pyautogui.PAUSE = 0.8
pyautogui.FAILSAFE = True

class Coord:
    send1 = (127, 87)
    addressbox = (710, 216)
    sgp = (715, 303)
    send2 = (77, 558)
    ok = (760, 463)
    metamask = (1251, 47)
    gasfrom = (1184, 306)
    gasto = (1193, 306)
    submit = (1110, 440)
    wallet = (1110, 440)

class Color:
    send1 = (133, 192, 218)
    gasfrom = (236, 236, 236)
    submit = (106, 195, 96)
    
def click(coordinates):
    pyautogui.click(coordinates[0], coordinates[1])
    
def matchcolor(coordinates, color):
    pyautogui.pixelMatchesColor(coordinates[0], coordinates[1], color)
    
def drag(coordinates):
    pyautogui.dragTo(coordinates[0], coordinates[1], duration = 0.5)
    
#excel data to list
li = []
sheet = openpyxl.load_workbook('twitterairdrop.xlsx')['Sheet1']
for i in range(100):
    d = {}
    x = 0 #update to row you wish to start iterating from
    d['user'] = sheet.cell(row=i+x, column=2).value
    d['address'] = sheet.cell(row=i+x, column=3).value
    d['tokens'] = sheet.cell(row=i+x, column=12).value
    li.append(d)
    
print("Program started!")
for person in li:
    while not matchcolor(Coord.send1, Color.send1):
        sleep(0.5)
    click(Coord.send1)
    print(f"Entering {person['user']}'s info...")
    click(Coord.addressbox)
    pyautogui.typewrite(person["address"] + "\t" + str(person["tokens"]))
    click(Coord.sgp)
    sleep(1)
    click(Coord.send2)
    sleep(1)
    click(Coord.ok)
    sleep(1)
    click(Coord.metamask)
    sleep(1)
    while not matchcolor(Coord.gasfrom, Color.gasfrom):
        sleep(0.5)
    click(Coord.gasfrom)
    drag(Coord.gasto)
    pyautogui.typewrite("5")    #common gas price
    while not matchcolor(Coord.submit, Color.submit):
        sleep(0.5)
    click(Coord.submit)
    sleep(1)
    click(Coord.wallet)
