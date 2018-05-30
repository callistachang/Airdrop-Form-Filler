import openpyxl
import pyautogui
from time import sleep

pyautogui.PAUSE = 0.8
pyautogui.FAILSAFE = True

#coordinates & rgb values
blueSend = (127, 87)
addressBox = (710, 216)
pickSGPay = (715, 303)
sendParticulars = (77, 558)
okButton = (760, 463)
clickMetamask = (1251, 47)
gasClickFrom = (1184, 306)
gasDragTo = (1193, 306)
greenSubmit = (1110, 440)
clickWallet = (73, 89)
blueSendColor = (133, 192, 218)
gasColor = (236, 236, 236)
greenSubmitColor = (106, 195, 96)

#excel data to list
li = []
sheet = openpyxl.load_workbook('facebookairdrop.xlsx')['Sheet1']
for i in range(9):
    d = {}
    x = 0 #update to row you wish to start iterating from
    d['user'] = sheet.cell(row=i+x, column=2).value
    d['address'] = sheet.cell(row=i+x, column=3).value
    d['tokens'] = sheet.cell(row=i+x, column=12).value
    li.append(d)

#the program!
print('WE HAVE BEGUN')
for person in li:
    while not pyautogui.pixelMatchesColor(blueSend[0], blueSend[1], blueSendColor):
        sleep(0.5)
    pyautogui.click(blueSend[0], blueSend[1])
    print("Entering %s's info..." % (person['user']))
    pyautogui.click(addressBox[0], addressBox[1])
    pyautogui.typewrite(person['address'] + '\t' + str(person['tokens']))
    pyautogui.click(pickSGPay[0], pickSGPay[1])
    sleep(2)
    pyautogui.click(sendParticulars[0], sendParticulars[1])
    sleep(1)
    pyautogui.click(okButton[0], okButton[1])
    sleep(1)
    pyautogui.click(clickMetamask[0], clickMetamask[1])
    '''
    while not pyautogui.pixelMatchesColor(gasClickFrom[0], gasClickFrom[1], gasColor):
        sleep(0.5)
    sleep(1)
    pyautogui.click(gasClickFrom[0], gasClickFrom[1])
    pyautogui.dragTo(gasDragTo[0], gasDragTo[1], duration=0.5)
    pyautogui.typewrite('5')
    '''
    while not pyautogui.pixelMatchesColor(greenSubmit[0], greenSubmit[1], greenSubmitColor):
        sleep(0.5)
    sleep(1)
    pyautogui.click(greenSubmit[0], greenSubmit[1])
    sleep(1)
    pyautogui.click(clickWallet[0], clickWallet[1])
    sleep(30)
